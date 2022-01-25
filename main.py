from flask import Flask, render_template, url_for, request, flash, abort, redirect
from functools import wraps
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView
from sqlalchemy.sql import func
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from os import path
from flask_login import LoginManager, login_required, current_user, login_user, logout_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

db = SQLAlchemy()
DB_NAME = "database.db"

app = Flask(__name__)
app.config['SECRET_KEY'] = 'hjshjhdjah kjshkjdhjs'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_NAME}'
app.config['FLASK_ADMIN_SWATCH'] = 'cerulean'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
UPLOAD_FOLDER = './static/upload'
ALLOWED_EXTENSIONS = {'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


class Week(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.Integer, default=0)
    date = db.Column(db.DateTime(timezone=True), default=func.now())


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subj_name = db.Column(db.String(50))
    attend_percent = db.Column(db.Float)
    date = db.Column(db.DateTime(timezone=True), default=func.now())
    week = db.Column(db.Integer)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True, autoincrement=False)
    first_name = db.Column(db.String(150))
    group = db.Column(db.String(150))
    password = db.Column(db.String(150))
    role = db.Column(db.String(150))
    attendance = db.relationship('Attendance')

    def is_admin(self):
        return self.role == 'admin'

    def is_user(self):
        return self.role == 'user'


admin = Admin(app, name='app', template_mode='bootstrap3')
admin.add_view(ModelView(User, db.session))
admin.add_view(ModelView(Attendance, db.session))
admin.add_view(ModelView(Week, db.session))

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)


@login_manager.user_loader
def load_user(id):
    return User.query.get(int(id))


def superuser(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = current_user
        if not user.is_authenticated:
            abort(404)
        if not user.is_admin():
            abort(404)
        return f(*args, **kwargs)

    return decorated_function


@app.before_first_request
def restrict_admin_url():
    endpoint = 'admin.index'
    url = url_for(endpoint)
    admin_index = app.view_functions.pop(endpoint)

    @app.route(url, endpoint=endpoint)
    @superuser
    def secure_admin_index():
        return admin_index()


@app.before_first_request
def default_admin():
    user = User.query.filter_by(first_name="root").first()
    if user:
        pass
    else:
        new_user = User(id=0000, first_name="root", password=generate_password_hash("root", method='sha256'), role="admin")
        db.session.add(new_user)
        db.session.commit()


@app.route('/update_default_admin', methods=['GET', 'POST'])
@login_required
@superuser
def update_default_admin():
    if request.method == 'POST':
        idc = request.form.get('code')
        password1 = request.form.get('password1')
        password2 = request.form.get('password2')
        if len(idc) < 4:
            flash('code must be greater than 3 characters.', category='error')
        elif password1 != password2:
            flash('Passwords don\'t match.', category='error')
        elif len(password1) < 7:
            flash('Password must be at least 7 characters.', category='error')
        else:
            try:
                df_admin = User.query.filter_by(first_name='root').first()
                df_admin.id = idc
                df_admin.password = generate_password_hash(password1, method='sha256')
                db.session.commit()
                login_user(df_admin, remember=True)
            except IntegrityError:
                db.session.rollback()
                flash('database error!', category='error')
            flash('Account created!', category='success')
            return redirect(url_for('setting'))

    return render_template("update-df-admin.html", user=current_user)


@app.route('/')
@login_required
def home():
    if current_user.is_admin():
        return redirect(url_for('setting'))
    else:
        q = db.session.query(Attendance).join(User).filter(User.id == current_user.id, Attendance.attend_percent.isnot(None)).order_by(Attendance.week.desc()).all()
        n = q[0].week
        lst = [[] for x in range(n)]
        for i in range(n):
            for element in q:
                if element.week == i + 1:
                    lst[i].append(element)
        return render_template("home.html", user=current_user, attendance=lst)


@app.route('/all_attendance')
@login_required
def all_attendance():
    if current_user.is_admin():
        return redirect(url_for('setting'))
    else:
        q = db.session.query(Attendance).join(User).filter(User.id == current_user.id, Attendance.attend_percent.isnot(None)).order_by(Attendance.week.desc()).all()
        n = q[0].week
        lst = [[] for x in range(n)]
        for i in range(n):
            for element in q:
                if element.week == i + 1:
                    lst[i].append(element)
        return render_template("all-attendance.html", user=current_user, attendance=lst)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        acid = request.form.get('acid')
        user = User.query.filter_by(id=acid).first()
        password = "st" + str(request.form.get('acid'))
        if user:
            if user.is_admin():
                return redirect(url_for('login_admin'))
            elif user.password == password:
                flash('Logged in successfully!', category='success')
                login_user(user, remember=True)
                return redirect(url_for('home'))
            else:
                flash('Incorrect password, try again.', category='error')
        else:
            flash('Id does not exist.', category='error')

    return render_template("login.html", user=current_user)


@app.route('/login-admin', methods=['GET', 'POST'])
def login_admin():
    if request.method == 'POST':
        acid = request.form.get('acid')
        password = request.form.get('password')

        user = User.query.filter_by(id=acid).first()
        if user:
            if check_password_hash(user.password, password):
                flash('Logged in successfully!', category='success')
                login_user(user, remember=True)
                return redirect(url_for('setting'))
            else:
                flash('Incorrect password, try again.', category='error')
        else:
            flash('user does not exist.', category='error')

    return render_template("login-admin.html", user=current_user)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/create-admin', methods=['GET', 'POST'])
@superuser
def create_admin():
    if request.method == 'POST':
        idc = request.form.get('code')
        first_name = request.form.get('firstName')
        password1 = request.form.get('password1')
        password2 = request.form.get('password2')
        role = "admin"

        user = User.query.filter_by(id=idc).first()
        if user:
            flash('code already exists.', category='error')
        elif len(idc) < 4:
            flash('code must be greater than 3 characters.', category='error')
        elif len(first_name) < 2:
            flash('First name must be greater than 1 character.', category='error')
        elif password1 != password2:
            flash('Passwords don\'t match.', category='error')
        elif len(password1) < 7:
            flash('Password must be at least 7 characters.', category='error')
        else:
            try:
                new_user = User(id=idc, first_name=first_name,
                                password=generate_password_hash(password1, method='sha256'), role=role)
                db.session.add(new_user)
                db.session.commit()
                login_user(new_user, remember=True)
            except IntegrityError:
                db.session.rollback()
                flash('database error!', category='error')
            flash('Account created!', category='success')
            return redirect(url_for('setting'))

    return render_template("sign_up.html", user=current_user)


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/insert-student', methods=['GET', 'POST'])
@superuser
def insert_student():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(path.join(app.config['UPLOAD_FOLDER'], filename))
            wb = load_workbook('./static/upload/' + filename)
            ws = wb["Sheet1"]
            try:
                for row in range(2, ws.max_row + 1):
                    idc = ws.cell(row=row, column=2).value
                    first_name = ws.cell(row=row, column=3).value
                    group = ws.cell(row=row, column=1).value
                    password = "st" + str(idc)
                    role = "user"
                    new_user = User(id=idc, first_name=first_name, group=group, password=password, role=role)
                    db.session.add(new_user)
                db.session.commit()
                flash('success insert', category='success')

            except IntegrityError:
                db.session.rollback()
                flash('error in data base', category='error')
            return redirect(url_for('setting'))

    return render_template('upload-file.html', user=current_user)


@app.route('/insert-attendance', methods=['GET', 'POST'])
@superuser
def insert_attendance():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(path.join(app.config['UPLOAD_FOLDER'], filename))
            wb = load_workbook('./static/upload/' + filename)
            last_week = Week.query.order_by(Week.date).first()
            if last_week:
                week = last_week.number + 1
            else:
                week = 1
            try:
                for ws in wb.worksheets:
                    for col in range(4, ws.max_column + 1):
                        for row in range(2, ws.max_row + 1):
                            attend = ws.cell(row=row, column=col).value
                            if type(attend) == float:
                                attend_per = round(attend * 100)
                            elif type(attend) == str:
                                attend_per = None
                            else:
                                attend_per = float(attend)
                            subj_name = ws.cell(row=1, column=col).value
                            stud_id = ws.cell(row=row, column=1).value
                            new_attendance = Attendance(subj_name=subj_name, attend_percent=attend_per, user_id=stud_id,
                                                        week=week)
                            db.session.add(new_attendance)
                new_week = Week(number=week)
                db.session.add(new_week)
                db.session.commit()
                flash('success insert', category='success')
            except IntegrityError:
                db.session.rollback()
                flash('error in data base', category='error')
            return redirect(url_for('setting'))
    return render_template('upload-attendance.html', user=current_user)


@app.route('/data', methods=['GET', 'POST'])
def data():
    return render_template('data.html', user=current_user)


@app.route('/setting', methods=['GET', 'POST'])
@superuser
def setting():
    return render_template('setting.html', user=current_user)


db.init_app(app)
if not path.exists('/' + DB_NAME):
    db.create_all(app=app)
    print('Created Database!')

if __name__ == '__main__':
    app.run(debug=True)
